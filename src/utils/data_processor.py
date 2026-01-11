"""
Data processing and export module.
Handles data validation, transformation, and Excel export.
"""

import pandas as pd
from datetime import datetime, timedelta
import os
from src.config.config import (
    EXCEL_OUTPUT_FILE, SHEET_CHANNEL_STATIC, SHEET_CHANNEL_EVOLVING,
    SHEET_VIDEOS_DATA, SHEET_SHORTS_DATA,
    CHANNEL_STATIC_FIELDS, CHANNEL_EVOLVING_FIELDS, VIDEO_FIELDS, DEBUG_MODE
)


def calculate_daily_metrics(channel_data):
    """
    Calculate daily metrics (subscriber change, view change, growth rate) by comparing
    with the most recent previous data from Evolving_Data sheet.
    
    Args:
        channel_data (dict): Current channel data
        
    Returns:
        dict: Channel data with calculated daily metrics
    """
    try:
        # Check if Excel file exists
        if not os.path.exists(EXCEL_OUTPUT_FILE):
            # First run, no historical data
            channel_data['daily_subscriber_change'] = 0
            channel_data['daily_views_change'] = 0
            channel_data['growth_rate'] = 0.0
            return channel_data
        
        # Read existing evolving data
        df_evolving = pd.read_excel(EXCEL_OUTPUT_FILE, sheet_name=SHEET_CHANNEL_EVOLVING)
        
        # Filter for this channel
        channel_id = channel_data.get('channel_id')
        if not channel_id:
            channel_data['daily_subscriber_change'] = 0
            channel_data['daily_views_change'] = 0
            channel_data['growth_rate'] = 0.0
            return channel_data
        
        channel_history = df_evolving[df_evolving['channel_id'] == channel_id]
        
        if len(channel_history) == 0:
            # First time tracking this channel
            channel_data['daily_subscriber_change'] = 0
            channel_data['daily_views_change'] = 0
            channel_data['growth_rate'] = 0.0
            return channel_data
        
        # Get most recent previous entry
        # Sort by scrape_date to get the latest
        channel_history['scrape_date'] = pd.to_datetime(channel_history['scrape_date'])
        channel_history = channel_history.sort_values('scrape_date', ascending=False)
        previous_data = channel_history.iloc[0]
        
        # Calculate changes
        current_subs = channel_data.get('subscribers', 0)
        previous_subs = previous_data.get('subscribers', 0)
        daily_sub_change = current_subs - previous_subs
        
        current_views = channel_data.get('total_views', 0)
        previous_views = previous_data.get('total_views', 0)
        daily_view_change = current_views - previous_views
        
        # Calculate growth rate (percentage change in subscribers)
        if previous_subs > 0:
            growth_rate = ((current_subs - previous_subs) / previous_subs) * 100
        else:
            growth_rate = 0.0
        
        # Update channel data
        channel_data['daily_subscriber_change'] = int(daily_sub_change)
        channel_data['daily_views_change'] = int(daily_view_change)
        channel_data['growth_rate'] = round(growth_rate, 2)
        
        if DEBUG_MODE:
            print(f"\n✓ Daily metrics calculated:")
            print(f"  - Subscriber change: {daily_sub_change:+,}")
            print(f"  - Views change: {daily_view_change:+,}")
            print(f"  - Growth rate: {growth_rate:+.2f}%")
        
        return channel_data
        
    except Exception as e:
        if DEBUG_MODE:
            print(f"  ⚠ Could not calculate daily metrics: {e}")
        channel_data['daily_subscriber_change'] = 0
        channel_data['daily_views_change'] = 0
        channel_data['growth_rate'] = 0.0
        return channel_data


def format_duration(seconds):
    """
    Convert duration from seconds to MM:SS format.
    
    Args:
        seconds (int or float): Duration in seconds
        
    Returns:
        str: Formatted duration as MM:SS
    """
    if not seconds or pd.isna(seconds):
        return "00:00"
    
    try:
        total_seconds = int(float(seconds))
        minutes = total_seconds // 60
        secs = total_seconds % 60
        return f"{minutes:02d}:{secs:02d}"
    except (ValueError, TypeError):
        return "00:00"


def validate_channel_data(channel_data):
    """
    Validate channel data completeness.
    
    Args:
        channel_data (dict): Channel data dictionary
        
    Returns:
        dict: Validated data
    """
    if not channel_data:
        return None
    
    # Check required fields
    required_fields = ['channel_handle', 'subscribers', 'total_views']
    for field in required_fields:
        if field not in channel_data or channel_data[field] is None:
            print(f"✗ Missing required field: {field}")
            return None
    
    if DEBUG_MODE:
        print("✓ Channel data validation passed")
    
    return channel_data


def validate_video_data(videos):
    """
    Validate video data list.
    
    Args:
        videos (list): List of video data dictionaries
        
    Returns:
        list: Validated videos
    """
    if not videos:
        return []
    
    valid_videos = []
    required_fields = ['url', 'title', 'view_count']
    
    for video in videos:
        # Check required fields
        if all(field in video and video[field] is not None for field in required_fields):
            valid_videos.append(video)
        else:
            if DEBUG_MODE:
                print(f"✗ Skipping invalid video: {video.get('title', 'Unknown')}")
    
    if DEBUG_MODE:
        print(f"✓ Video data validation: {len(valid_videos)}/{len(videos)} valid")
    
    return valid_videos


def process_video_data(basic_videos, detailed_data_list, channel_id):
    """
    Merge basic and detailed video data.
    
    Args:
        basic_videos (list): Basic video info from list page
        detailed_data_list (list): Detailed video info from individual pages
        channel_id (str): Channel ID to link videos to channel
        
    Returns:
        list: Combined video data
    """
    videos = []
    
    for i, basic_video in enumerate(basic_videos):
        video_data = {**basic_video}
        
        # Add channel_id for linking
        video_data['channel_id'] = channel_id
        
        if i < len(detailed_data_list):
            detailed = detailed_data_list[i]
            # Update with detailed data, but preserve view_count if detailed has better value
            for key, value in detailed.items():
                if key == 'view_count':
                    # Use detailed view_count if it's greater (for shorts from detailed scraping)
                    if value and value > video_data.get('view_count', 0):
                        video_data[key] = value
                elif key == 'upload_date':
                    # Use detailed upload_date if basic doesn't have it or detailed is better
                    if value and (not video_data.get('upload_date') or value):
                        video_data[key] = value
                else:
                    # For all other fields, use detailed data if available
                    if value is not None and value != '':
                        video_data[key] = value
        
        # Determine if short - refined with actual duration
        duration = video_data.get('duration', 0)
        if isinstance(duration, str):
            # Convert MM:SS to seconds if needed
            try:
                if ':' in duration:
                    parts = duration.split(':')
                    duration = int(parts[0]) * 60 + int(parts[1])
                else:
                    duration = int(duration)
            except:
                duration = 0
        
        is_short = duration < 60 or '#shorts' in video_data.get('title', '').lower() or video_data.get('is_short', False)
        video_data['is_short'] = is_short
        
        # Format duration to MM:SS
        video_data['duration'] = format_duration(duration)
        
        videos.append(video_data)
    
    return videos


def calculate_channel_metrics(channel_data, long_videos, shorts):
    """
    Calculate aggregate channel metrics from videos and shorts.
    
    Args:
        channel_data (dict): Channel data
        long_videos (list): List of long-form video data
        shorts (list): List of shorts data
        
    Returns:
        dict: Updated channel data with calculated metrics
    """
    if not long_videos and not shorts:
        return channel_data
    
    try:
        # Only update counts if they weren't already set from actual YouTube counts
        # This preserves the actual total counts from YouTube tabs
        if 'video_count' not in channel_data or channel_data['video_count'] == 0:
            channel_data['video_count'] = len(long_videos)
        if 'shorts_count' not in channel_data or channel_data['shorts_count'] == 0:
            channel_data['shorts_count'] = len(shorts)
        if 'total_content_count' not in channel_data or channel_data['total_content_count'] == 0:
            channel_data['total_content_count'] = channel_data.get('video_count', 0) + channel_data.get('shorts_count', 0)
        
        # Get last posted date from all content
        all_content = long_videos + shorts
        dates = [v.get('upload_date', '') for v in all_content if v.get('upload_date', '')]
        if dates:
            channel_data['last_posted_date'] = dates[0]  # First item is latest
        
        if DEBUG_MODE:
            print(f"✓ Channel metrics calculated")
            print(f"  - Long-form videos (actual): {channel_data.get('video_count', 0)}")
            print(f"  - Shorts (actual): {channel_data.get('shorts_count', 0)}")
            print(f"  - Total content: {channel_data.get('total_content_count', 0)}")
            print(f"  - Extracted videos: {len(long_videos)}")
            print(f"  - Extracted shorts: {len(shorts)}")
        
        return channel_data
    
    except Exception as e:
        print(f"✗ Error calculating metrics: {e}")
        return channel_data


def export_to_excel(channel_data, videos, shorts):
    """
    Export channel data (static + evolving), videos, and shorts to Excel file with separate sheets.
    Follows the 3-part structure defined in the PDF:
    - Static_Data: Channel-level static info (rarely changes)
    - Evolving_Data: Channel-level metrics updated daily
    - Videos_Data: Long-form videos
    - Shorts_Data: Short-form videos
    
    Args:
        channel_data (dict): Channel data
        videos (list): List of long-form video data
        shorts (list): List of shorts data
        
    Returns:
        bool: True if successful, False otherwise
    """
    try:
        import os
        from openpyxl import load_workbook
        
        # Get absolute path
        abs_path = os.path.abspath(EXCEL_OUTPUT_FILE)
        
        # Ensure data directory exists
        os.makedirs(os.path.dirname(abs_path), exist_ok=True)
        
        # Prepare STATIC data (rarely changing channel info)
        static_data = {k: channel_data.get(k, '') for k in CHANNEL_STATIC_FIELDS if k in channel_data}
        df_static_new = pd.DataFrame([static_data])
        
        # Prepare EVOLVING data (daily updated metrics)
        evolving_data = {k: channel_data.get(k, '') for k in CHANNEL_EVOLVING_FIELDS if k in channel_data}
        df_evolving_new = pd.DataFrame([evolving_data])
        
        # Check if file exists to append or create new
        file_exists = os.path.exists(abs_path)
        
        if file_exists:
            # Append to existing sheets
            
            # STATIC DATA - Check if channel already exists, update if so
            try:
                df_static_existing = pd.read_excel(abs_path, sheet_name=SHEET_CHANNEL_STATIC)
                channel_id = channel_data.get('channel_id')
                
                if channel_id and channel_id in df_static_existing['channel_id'].values:
                    # Update existing channel
                    df_static_existing.loc[df_static_existing['channel_id'] == channel_id] = df_static_new.iloc[0]
                    df_static = df_static_existing
                else:
                    # Add new channel
                    df_static = pd.concat([df_static_existing, df_static_new], ignore_index=True)
            except:
                df_static = df_static_new
            
            # EVOLVING DATA - Always append (daily snapshots)
            try:
                df_evolving_existing = pd.read_excel(abs_path, sheet_name=SHEET_CHANNEL_EVOLVING)
                df_evolving = pd.concat([df_evolving_existing, df_evolving_new], ignore_index=True)
            except:
                df_evolving = df_evolving_new
            
            # Write both sheets
            book = load_workbook(abs_path)
            with pd.ExcelWriter(abs_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df_static.to_excel(writer, sheet_name=SHEET_CHANNEL_STATIC, index=False)
                df_evolving.to_excel(writer, sheet_name=SHEET_CHANNEL_EVOLVING, index=False)
        else:
            # Create new file
            with pd.ExcelWriter(abs_path, engine='openpyxl') as writer:
                df_static_new.to_excel(writer, sheet_name=SHEET_CHANNEL_STATIC, index=False)
                df_evolving_new.to_excel(writer, sheet_name=SHEET_CHANNEL_EVOLVING, index=False)
        
        if DEBUG_MODE:
            print(f"[OK] Channel static data saved")
            print(f"[OK] Channel evolving data saved")
        
        # VIDEOS data sheet - append logic
        if videos:
            df_videos_new = pd.DataFrame(videos)
            available_cols = [col for col in VIDEO_FIELDS if col in df_videos_new.columns]
            df_videos_new = df_videos_new[available_cols]
            
            if file_exists:
                try:
                    df_videos_existing = pd.read_excel(abs_path, sheet_name=SHEET_VIDEOS_DATA)
                    df_videos = pd.concat([df_videos_existing, df_videos_new], ignore_index=True)
                except:
                    df_videos = df_videos_new
                
                book = load_workbook(abs_path)
                with pd.ExcelWriter(abs_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    df_videos.to_excel(writer, sheet_name=SHEET_VIDEOS_DATA, index=False)
            else:
                with pd.ExcelWriter(abs_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    df_videos_new.to_excel(writer, sheet_name=SHEET_VIDEOS_DATA, index=False)
            
            if DEBUG_MODE:
                total_videos = len(df_videos) if file_exists and 'df_videos' in locals() else len(df_videos_new)
                print(f"[OK] Videos data sheet updated ({len(df_videos_new)} new, total: {total_videos})")
        
        # SHORTS data sheet - append logic
        if shorts:
            df_shorts_new = pd.DataFrame(shorts)
            available_cols = [col for col in VIDEO_FIELDS if col in df_shorts_new.columns]
            df_shorts_new = df_shorts_new[available_cols]
            
            if file_exists:
                try:
                    df_shorts_existing = pd.read_excel(abs_path, sheet_name=SHEET_SHORTS_DATA)
                    df_shorts = pd.concat([df_shorts_existing, df_shorts_new], ignore_index=True)
                except:
                    df_shorts = df_shorts_new
                
                book = load_workbook(abs_path)
                with pd.ExcelWriter(abs_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    df_shorts.to_excel(writer, sheet_name=SHEET_SHORTS_DATA, index=False)
            else:
                with pd.ExcelWriter(abs_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    df_shorts_new.to_excel(writer, sheet_name=SHEET_SHORTS_DATA, index=False)
            
            if DEBUG_MODE:
                total_shorts = len(df_shorts) if file_exists and 'df_shorts' in locals() else len(df_shorts_new)
                print(f"[OK] Shorts data sheet updated ({len(df_shorts_new)} new, total: {total_shorts})")
        
        # Verify file was created
        if os.path.exists(abs_path):
            file_size = os.path.getsize(abs_path)
            print(f"[OK] Saved to: {abs_path}")
            print(f"[OK] File size: {file_size:,} bytes")
        else:
            print(f"[WARNING] File may not have been saved properly")
        
        return True
    
    except Exception as e:
        print(f"[ERROR] Error exporting to Excel: {e}")
        import traceback
        traceback.print_exc()
        return False


def load_from_excel(filename=EXCEL_OUTPUT_FILE):
    """
    Load data from Excel file.
    
    Args:
        filename (str): Excel file name
        
    Returns:
        tuple: (channel_data_df, videos_data_df) or (None, None) if error
    """
    try:
        df_channel = pd.read_excel(filename, sheet_name=SHEET_CHANNEL_DATA)
        # Read videos and keep duration as string to preserve MM:SS format
        df_videos = pd.read_excel(filename, sheet_name=SHEET_VIDEOS_DATA, dtype={'duration': str})
        # Also read shorts if sheet exists
        try:
            df_shorts = pd.read_excel(filename, sheet_name='Shorts_Data', dtype={'duration': str})
        except:
            df_shorts = pd.DataFrame()
        
        if DEBUG_MODE:
            print(f"✓ Loaded {len(df_channel)} channels and {len(df_videos)} videos")
        
        return df_channel, df_videos
    
    except Exception as e:
        print(f"✗ Error loading Excel file: {e}")
        return None, None
